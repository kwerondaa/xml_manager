import openpyxl as xl

wbook = xl.load_workbook('C:\\Users\\Kweronda\\Documents\\python\\files\\excel\\transactions.xlsx')
sheet = wbook['Sheet1']


for row in range(2, sheet.max_row+1):
    cell =sheet.cell(row, 3)
    corrected_price =cell.value * 0.9

    corrected_price_cell = sheet.cell(row, 4) 
    corrected_price_cell.value =corrected_price

wbook.save('transactions2.xlsx')

